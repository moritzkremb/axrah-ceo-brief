from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

wb = Workbook()
ws = wb.active
ws.title = "ROI Calculator"

# ── HELPERS ──────────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(color="D0D8E4", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color="D0D8E4", style="thin"):
    s = Side(style=style, color=color)
    n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=s)

def font(bold=False, size=10, color="1A2B40", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(ws, cell_ref, value=None, bold=False, size=10,
               color="1A2B40", bg=None, align_h="left",
               number_format=None, italic=False, wrap=False):
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(align_h, "center", wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if number_format:
        c.number_format = number_format
    return c

# ── COLUMN / ROW DIMENSIONS ───────────────────────────────────────────────────
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 5
ws.column_dimensions["E"].width = 34
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 2

for r in range(1, 60):
    ws.row_dimensions[r].height = 18

# ── PALETTE ───────────────────────────────────────────────────────────────────
RED      = "C8102E"
NAVY     = "0F1B2D"
NAVY_MID = "1A2B40"
NAVY_L   = "243449"
INPUT_BG = "FFFBEF"   # warm cream  — editable
CALC_BG  = "F0F6FF"   # pale blue   — calculated
SECT_BG  = "1A2B40"   # navy        — section headers
WHITE    = "FFFFFF"
GREEN    = "007A4C"
GREEN_BG = "E6F7F0"
RED_BG   = "FFF0F2"
MUTED    = "6B7C93"
BORDER_C = "D0D8E4"

# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════
ws.merge_cells("B1:F1")
c = ws["B1"]
c.value = "AXRAH  —  Investment ROI Calculator"
c.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
ws.row_dimensions[1].height = 36

ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = ("Enter your inputs in the yellow cells. All other cells calculate automatically.  "
           "|  Confidential — for client use only")
c.font = Font(size=9, color="8EA0B4", italic=True, name="Calibri")
c.fill = fill(NAVY_MID)
c.alignment = align("left", "center")
ws.row_dimensions[2].height = 22

ws.row_dimensions[3].height = 10

# ══════════════════════════════════════════════════════════════════════════════
# LEFT COLUMN — INPUTS  (col B label, col C value)
# ══════════════════════════════════════════════════════════════════════════════

def section_header_left(row, text):
    ws.merge_cells(f"B{row}:C{row}")
    c = ws[f"B{row}"]
    c.value = text
    c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
    c.fill = fill(SECT_BG)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 20

def input_row(row, label, cell_ref, default, fmt=None, note=""):
    # Label
    lc = ws[f"B{row}"]
    lc.value = label
    lc.font = font(size=9, color=NAVY_MID)
    lc.fill = fill("F7F9FC")
    lc.alignment = align("left", "center")
    lc.border = border(BORDER_C)

    # Input cell
    ic = ws[cell_ref]
    ic.value = default
    ic.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
    ic.fill = fill(INPUT_BG)
    ic.alignment = align("center", "center")
    ic.border = Border(
        left=Side(style="medium", color=RED),
        right=Side(style="thin", color=BORDER_C),
        top=Side(style="thin", color=BORDER_C),
        bottom=Side(style="thin", color=BORDER_C),
    )
    if fmt:
        ic.number_format = fmt

# ── DEVICE SELECTOR ──────────────────────────────────────────────────────────
section_header_left(4, "  DEVICE SELECTION")

lc = ws["B5"]
lc.value = "Select device"
lc.font = font(size=9, color=NAVY_MID)
lc.fill = fill("F7F9FC")
lc.alignment = align("left", "center")
lc.border = border(BORDER_C)

ic = ws["C5"]
ic.value = "Chamber Ultra — $26,999"
ic.font = Font(bold=True, size=10, color=RED, name="Calibri")
ic.fill = fill(INPUT_BG)
ic.alignment = align("center", "center")
ic.border = Border(
    left=Side(style="medium", color=RED),
    right=Side(style="thin", color=BORDER_C),
    top=Side(style="thin", color=BORDER_C),
    bottom=Side(style="thin", color=BORDER_C),
)

dv = DataValidation(
    type="list",
    formula1='"Chamber — $16,999,Chamber Ultra — $26,999"',
    allow_blank=False,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Invalid selection",
    error="Please choose from the dropdown list.",
)
ws.add_data_validation(dv)
dv.add("C5")

# Hidden device cost cell (used by formulas)
ws["C6"].value = '=IF(C5="Chamber — $16,999",16999,26999)'
ws["C6"].font = Font(size=9, color=WHITE, name="Calibri")
ws["C6"].fill = fill(WHITE)
ws["C6"].number_format = '"$"#,##0'
ws.row_dimensions[6].height = 4

ws.row_dimensions[7].height = 8

# ── SESSION INPUTS ────────────────────────────────────────────────────────────
section_header_left(8, "  YOUR INPUTS  —  change any yellow cell")

input_row(9,  "Price per session ($)",          "C9",  75,   '"$"#,##0.00')
input_row(10, "Monthly operating costs ($)",    "C10", 200,  '"$"#,##0')
input_row(11, "Sessions per day",               "C11", 6,    '0')
input_row(12, "Days open per week",             "C12", 5,    '0')
input_row(13, "Occupancy rate (%)",             "C13", 0.65, '0%')
input_row(14, "Avg sessions per client / month","C14", 4,    '0')

ws.row_dimensions[15].height = 8

# ── DATA VALIDATION on inputs ─────────────────────────────────────────────────
for cell, mn, mx, msg in [
    ("C9",  1,   500,  "Enter a session price between $1 and $500"),
    ("C10", 0,   50000,"Enter monthly operating costs"),
    ("C11", 1,   30,   "Enter sessions per day (1–30)"),
    ("C12", 1,   7,    "Enter days per week (1–7)"),
    ("C13", 0.05,1.0,  "Enter occupancy as a decimal, e.g. 0.65 for 65%"),
    ("C14", 1,   30,   "Enter average sessions per client per month"),
]:
    v = DataValidation(type="decimal", operator="between",
                       formula1=str(mn), formula2=str(mx),
                       allow_blank=False,
                       showErrorMessage=True,
                       errorTitle="Out of range", error=msg)
    ws.add_data_validation(v)
    v.add(cell)

# ── KEY METRIC LEGEND ─────────────────────────────────────────────────────────
ws.row_dimensions[16].height = 8
section_header_left(17, "  LEGEND")

for row, bg, txt in [
    (18, INPUT_BG, "  Yellow = input cell — edit freely"),
    (19, CALC_BG,  "  Blue   = calculated automatically"),
    (20, GREEN_BG, "  Green  = profit / positive return"),
]:
    ws.merge_cells(f"B{row}:C{row}")
    c = ws[f"B{row}"]
    c.value = txt
    c.font = Font(size=9, color=NAVY_MID, name="Calibri")
    c.fill = fill(bg)
    c.alignment = align("left", "center")
    c.border = border(BORDER_C)
    ws.row_dimensions[row].height = 17

# ══════════════════════════════════════════════════════════════════════════════
# RIGHT COLUMN — RESULTS  (col E label, col F value)
# ══════════════════════════════════════════════════════════════════════════════

def section_header_right(row, text):
    ws.merge_cells(f"E{row}:F{row}")
    c = ws[f"E{row}"]
    c.value = text
    c.font = Font(bold=True, size=9, color=WHITE, name="Calibri")
    c.fill = fill(RED)
    c.alignment = align("left", "center")
    ws.row_dimensions[row].height = 20

def result_row(row, label, formula, fmt, bg=CALC_BG, bold_val=False, note=""):
    lc = ws[f"E{row}"]
    lc.value = label
    lc.font = font(size=9, color=NAVY_MID)
    lc.fill = fill("F7F9FC")
    lc.alignment = align("left", "center")
    lc.border = border(BORDER_C)

    vc = ws[f"F{row}"]
    vc.value = formula
    vc.font = Font(bold=bold_val, size=11 if bold_val else 10, color=NAVY, name="Calibri")
    vc.fill = fill(bg)
    vc.alignment = align("center", "center")
    vc.border = border(BORDER_C)
    vc.number_format = fmt

# ── MONTHLY METRICS ───────────────────────────────────────────────────────────
section_header_right(4, "  MONTHLY METRICS")

result_row(5,  "Sessions per month",
           "=ROUND(C11*C13*C12*4.33,0)", '0', bold_val=False)
result_row(6,  "Monthly revenue",
           "=F5*C9",                       '"$"#,##0', bold_val=True)
result_row(7,  "Monthly operating costs",
           "=C10",                         '"$"#,##0')
result_row(8,  "Monthly profit",
           "=F6-F7",                       '"$"#,##0', bg=GREEN_BG, bold_val=True)
result_row(9,  "Clients served / month",
           "=IFERROR(ROUND(F5/C14,0),0)",  '0')

ws.row_dimensions[10].height = 8

# ── PAYBACK ───────────────────────────────────────────────────────────────────
section_header_right(11, "  PAYBACK PERIOD")

lc = ws["E12"]
lc.value = "Device cost"
lc.font = font(size=9, color=NAVY_MID)
lc.fill = fill("F7F9FC")
lc.alignment = align("left", "center")
lc.border = border(BORDER_C)

vc = ws["F12"]
vc.value = "=C6"
vc.font = Font(size=10, color=NAVY, name="Calibri")
vc.fill = fill(CALC_BG)
vc.alignment = align("center", "center")
vc.border = border(BORDER_C)
vc.number_format = '"$"#,##0'

lc2 = ws["E13"]
lc2.value = "Payback period (months)"
lc2.font = Font(bold=True, size=9, color=NAVY_MID, name="Calibri")
lc2.fill = fill("F7F9FC")
lc2.alignment = align("left", "center")
lc2.border = border(BORDER_C)

vc2 = ws["F13"]
vc2.value = '=IFERROR(IF(F8>0,CEILING(C6/F8,0.5),"N/A — check inputs"),"N/A")'
vc2.font = Font(bold=True, size=13, color=RED, name="Calibri")
vc2.fill = fill(RED_BG)
vc2.alignment = align("center", "center")
vc2.border = Border(
    left=Side(style="medium", color=RED),
    right=Side(style="medium", color=RED),
    top=Side(style="medium", color=RED),
    bottom=Side(style="medium", color=RED),
)
vc2.number_format = '0.0 "months"'

ws.row_dimensions[14].height = 8

# ── ANNUAL ────────────────────────────────────────────────────────────────────
section_header_right(15, "  ANNUAL METRICS")

result_row(16, "Annual revenue",           "=F6*12",       '"$"#,##0', bold_val=True)
result_row(17, "Annual operating costs",   "=F7*12",       '"$"#,##0')
result_row(18, "Annual gross profit",      "=F16-F17",     '"$"#,##0', bg=GREEN_BG, bold_val=True)

ws.row_dimensions[19].height = 8

# ── 3-YEAR PROJECTION ─────────────────────────────────────────────────────────
section_header_right(20, "  3-YEAR PROJECTION  (cumulative revenue)")

result_row(21, "Year 1 — revenue",         "=F16",         '"$"#,##0')
result_row(22, "Year 2 — revenue",         "=F16*2",       '"$"#,##0')
result_row(23, "Year 3 — revenue",         "=F16*3",       '"$"#,##0', bold_val=True)

ws.row_dimensions[24].height = 8

section_header_right(25, "  NET PROFIT AFTER DEVICE COST")

result_row(26, "Year 1 net profit",
           "=F18-C6",    '"$"#,##0', bg=CALC_BG)
result_row(27, "Year 2 net profit",
           "=F18*2-C6",  '"$"#,##0', bg=GREEN_BG, bold_val=True)
result_row(28, "Year 3 net profit",
           "=F18*3-C6",  '"$"#,##0', bg=GREEN_BG, bold_val=True)

# Colour negative Year 1 net profit red
from openpyxl.formatting.rule import CellIsRule
ws.conditional_formatting.add(
    "F26",
    CellIsRule(operator="lessThan", formula=["0"],
               font=Font(color="C8102E", bold=True, name="Calibri"),
               fill=fill(RED_BG))
)
ws.conditional_formatting.add(
    "F26:F28",
    CellIsRule(operator="greaterThanOrEqual", formula=["0"],
               font=Font(color="007A4C", bold=True, name="Calibri"),
               fill=fill(GREEN_BG))
)

# ── PAYBACK conditional ───────────────────────────────────────────────────────
ws.conditional_formatting.add(
    "F13",
    CellIsRule(operator="lessThanOrEqual", formula=["12"],
               font=Font(color="007A4C", bold=True, size=13, name="Calibri"),
               fill=fill(GREEN_BG))
)
ws.conditional_formatting.add(
    "F13",
    CellIsRule(operator="greaterThan", formula=["24"],
               font=Font(color="C8102E", bold=True, size=13, name="Calibri"),
               fill=fill(RED_BG))
)

# ── CHART ─────────────────────────────────────────────────────────────────────
# Build a small data table for the chart (rows 31-34, hidden area)
ws["E31"].value = "Year"
ws["F31"].value = "Cumulative Revenue"
ws["E32"].value = "Year 1"; ws["F32"].value = "=F16"
ws["E33"].value = "Year 2"; ws["F33"].value = "=F16*2"
ws["E34"].value = "Year 3"; ws["F34"].value = "=F16*3"
for r in range(31, 35):
    ws.row_dimensions[r].height = 0  # hidden

chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "3-Year Revenue Projection"
chart.y_axis.title = "Revenue ($)"
chart.x_axis.title = "Year"
chart.style = 2
chart.width = 12
chart.height = 8

data_ref = Reference(ws, min_col=6, min_row=31, max_row=34)
cats_ref = Reference(ws, min_col=5, min_row=32, max_row=34)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "C8102E"
chart.series[0].graphicalProperties.line.solidFill = "9E0C24"

ws.add_chart(chart, "B22")

# ══════════════════════════════════════════════════════════════════════════════
# FOOTNOTE / DISCLAIMER
# ══════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[37].height = 8
ws.merge_cells("B38:F38")
c = ws["B38"]
c.value = ("Formula: Monthly revenue = sessions/day × occupancy × days/week × 4.33 × session price.  "
           "Profit = revenue − operating costs.  Payback = device cost ÷ monthly profit.  "
           "Net profit deducts full device cost in Year 1.  "
           "All figures are estimates for planning purposes only. Actual results will vary.  "
           "© AXRAH 2026")
c.font = Font(size=8, color=MUTED, italic=True, name="Calibri")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[38].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# FREEZE PANES & SHEET PROTECTION
# ══════════════════════════════════════════════════════════════════════════════
ws.freeze_panes = "B4"

# Protect sheet but allow editing of input cells
ws.protection.sheet = True
ws.protection.password = ""  # no password — just visual lock indicator
ws.protection.enable()

# Unlock input cells so they remain editable
for cell_ref in ["C5", "C9", "C10", "C11", "C12", "C13", "C14"]:
    ws[cell_ref].protection = __import__('openpyxl.styles.protection',
                                          fromlist=['Protection']).Protection(locked=False)

# ══════════════════════════════════════════════════════════════════════════════
# PRINT SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_area = "A1:G40"

out_path = "/home/user/axrah-ceo-brief/sales/AXRAH-ROI-Calculator.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
